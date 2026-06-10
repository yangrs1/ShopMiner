"""生成测试用例 Excel 文件"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import yaml
import os

YAML_DIR = os.path.join(os.path.dirname(__file__), "data", "yaml")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "report", "test_cases.xlsx")

MODULE_MAP = {
    "auth.yaml": "认证模块",
    "product.yaml": "商品模块",
    "cart.yaml": "购物车模块",
    "order.yaml": "订单模块",
    "admin.yaml": "管理员模块",
    "analytics.yaml": "数据分析模块",
    "web.yaml": "Web UI",
}

PRIORITY_MAP = {
    "认证模块": {"register_success": "P0", "register_missing_fields": "P1", "register_duplicate": "P0",
                "login_success": "P0", "login_failure": "P0", "get_current_user": "P0"},
    "商品模块": {"get_products": "P0", "get_categories": "P1", "get_product_detail": "P0",
                "create_product": "P0", "update_product": "P1", "delete_product": "P1"},
    "购物车模块": {"add_to_cart": "P0", "get_cart": "P0", "update_cart": "P1",
                  "remove_cart_item": "P1", "clear_cart": "P2"},
    "订单模块": {"create_order": "P0", "get_orders": "P0", "get_order_detail": "P1",
                "cancel_order": "P0", "get_order_status_logs": "P2"},
    "管理员模块": {"get_all_orders": "P0", "ship_order": "P0", "deliver_order": "P0",
                  "refund_order": "P0", "get_all_users": "P1", "adjust_balance": "P1"},
    "数据分析模块": {"dashboard": "P0", "rfm_summary": "P0", "sales_trend": "P0", "sales_prediction": "P0",
                    "association_list": "P0", "association_for_product": "P0", "churn_list": "P0",
                    "churn_importance": "P1", "user_rfm": "P0", "user_trend": "P1",
                    "user_category_preference": "P1", "model_metrics": "P1", "hot_products": "P0",
                    "recompute": "P1", "last_compute_time": "P2", "churn_status_update": "P0"},
    "Web UI": {"register_flow": "P0", "login_flow": "P0", "browse_and_search": "P0",
               "cart_flow": "P0", "order_flow": "P0", "profile_flow": "P1", "admin_flow": "P0"},
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
P0_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
P1_FILL = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
P2_FILL = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
P3_FILL = PatternFill(start_color="C4C4C4", end_color="C4C4C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def get_priority_fill(priority):
    return {"P0": P0_FILL, "P1": P1_FILL, "P2": P2_FILL, "P3": P3_FILL}.get(priority, P3_FILL)


def load_all_yaml():
    all_cases = []
    for filename, module_name in MODULE_MAP.items():
        filepath = os.path.join(YAML_DIR, filename)
        if not os.path.exists(filepath):
            continue
        with open(filepath, encoding="utf-8") as f:
            data = yaml.safe_load(f)
        priority_map = PRIORITY_MAP.get(module_name, {})
        for group_name, cases in data.items():
            for idx, case in enumerate(cases, 1):
                priority = priority_map.get(group_name, "P2")
                tc_id = f"TC-{module_name[:2].upper()}-{idx:03d}"
                all_cases.append({
                    "id": tc_id,
                    "name": case.get("name", f"{group_name} #{idx}"),
                    "module": module_name,
                    "group": group_name,
                    "priority": priority,
                    "precondition": "用户已登录" if not case.get("no_auth") and module_name != "Web UI" else "无需登录" if case.get("no_auth") else "用户已登录",
                    "data": str({k: v for k, v in case.items() if k not in ("name", "expected_status", "expected_code")}),
                    "expected": f"HTTP {case.get('expected_status', 'N/A')}",
                })
    return all_cases


def generate_excel():
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    cases = load_all_yaml()
    wb = openpyxl.Workbook()

    # Sheet 1: 测试用例
    ws = wb.active
    ws.title = "测试用例"
    headers = ["编号", "名称", "模块", "分组", "优先级", "前置条件", "测试数据", "预期结果", "实际结果", "状态"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row, tc in enumerate(cases, 2):
        values = [tc["id"], tc["name"], tc["module"], tc["group"], tc["priority"],
                  tc["precondition"], tc["data"], tc["expected"], "", "未执行"]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            if col == 5:
                cell.fill = get_priority_fill(tc["priority"])
                cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    widths = [14, 30, 14, 24, 8, 14, 40, 14, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 2: 统计摘要
    ws2 = wb.create_sheet("统计摘要")
    summary_headers = ["模块", "P0", "P1", "P2", "P3", "合计"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    stats = {}
    for tc in cases:
        mod = tc["module"]
        pri = tc["priority"]
        stats.setdefault(mod, {"P0": 0, "P1": 0, "P2": 0, "P3": 0})
        stats[mod][pri] += 1

    row = 2
    totals = {"P0": 0, "P1": 0, "P2": 0, "P3": 0}
    for mod, pri_counts in sorted(stats.items()):
        ws2.cell(row=row, column=1, value=mod)
        for i, p in enumerate(["P0", "P1", "P2", "P3"], 2):
            ws2.cell(row=row, column=i, value=pri_counts[p])
            totals[p] += pri_counts[p]
        ws2.cell(row=row, column=6, value=sum(pri_counts.values()))
        row += 1

    ws2.cell(row=row, column=1, value="合计")
    for i, p in enumerate(["P0", "P1", "P2", "P3"], 2):
        ws2.cell(row=row, column=i, value=totals[p])
    ws2.cell(row=row, column=6, value=sum(totals.values()))

    for i in range(1, 7):
        ws2.column_dimensions[get_column_letter(i)].width = 14

    wb.save(OUTPUT_PATH)
    print(f"Excel generated: {OUTPUT_PATH}")
    print(f"Total test cases: {len(cases)}")
    print(f"P0: {totals['P0']}, P1: {totals['P1']}, P2: {totals['P2']}, P3: {totals['P3']}")


if __name__ == "__main__":
    generate_excel()
